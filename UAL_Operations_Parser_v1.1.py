"""
UAL Audit Log Parser
DFIR Tooling | Yuvi Kapoor
M365 Unified Audit Log – BEC Focused Parser
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import json
import csv
import os
import threading
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Colour Palette ──────────────────────────────────────────────────────────
DEEP_NAVY    = "#1A1D27"
NAVY_MID     = "#20243A"
NAVY_LIGHT   = "#252A40"
IRIS         = "#6374F8"
IRIS_HOVER   = "#7A8BFF"
IRIS_DIM     = "#3D4A9E"
TEXT_PRIMARY = "#E8EAF6"
TEXT_MUTED   = "#8890B0"
TEXT_DIM     = "#555D80"
SUCCESS      = "#4CAF82"
WARNING      = "#F59E0B"
DANGER       = "#EF4444"
BORDER       = "#2E3450"

# ─── Operation Classification ─────────────────────────────────────────────────
# FIX: Removed generic "Update" and "Move" from email_send — these fire on
# calendar items, contacts, tasks, and other non-email objects that have no
# Item.Subject structure. Kept only operations that reliably produce email
# AuditData with an Item block.
OP_EMAIL_SEND = {
    "Create", "Send", "SendAs", "SendOnBehalf",
    "MoveToFolder",       # explicit mailbox folder move (has Item block)
    "Copy",
}
OP_DELETE = {
    "HardDelete", "SoftDelete", "MoveToDeletedItems",
    "RecoverDeletedMessages", "PurgedMessages",
    "ApplyRecord",        # retention label applied — can precede purge
}
OP_MAIL_ACCESS = {"MailItemsAccessed"}
# FIX: Extended with real SharePoint/OneDrive ops seen in UAL exports.
# Sharing ops are high-value in BEC (data exfil via sharing links).
OP_FILE_FOLDER = {
    # Core file ops
    "FileModified", "FileAccessed", "FileUploaded", "FileDownloaded",
    "FilePreviewed", "FileDeleted", "FileMoved", "FileRenamed",
    "FileCheckedIn", "FileCheckedOut", "FileCheckOutDiscarded",
    "FileSyncDownloadedFull", "FileSyncUploadedFull",
    "FileVersionsAllDeleted", "FileVersionRecycled", "FileRestored",
    "FileRecycled",
    # Folder ops
    "FolderCreated", "FolderModified", "FolderDeleted",
    "FolderMoved", "FolderCopied", "FolderRecycled",
    # Sharing — critical for BEC data exfil detection
    "SharingInvitationCreated", "SharingInvitationAccepted",
    "SharingInvitationRevoked", "SharingInvitationBlocked",
    "AnonymousLinkCreated", "AnonymousLinkUpdated", "AnonymousLinkUsed",
    "AnonymousLinkRemoved",
    "CompanyLinkCreated", "CompanyLinkUsed", "CompanyLinkRemoved",
    "SecureLinkCreated", "SecureLinkUsed", "SecureLinkUpdated",
    "SecureLinkDeleted",
    "AddedToSecureLink", "RemovedFromSecureLink",
    "SharingSet", "SharingRevoked",
    "AccessRequestCreated", "AccessRequestApproved",
    # List/library ops
    "ListCreated", "ListDeleted", "ListUpdated",
    "ListItemCreated", "ListItemUpdated", "ListItemDeleted",
    "ListItemRecycled",
    "ListColumnAdded", "ListColumnUpdated", "ListColumnDeleted",
    # Group / membership
    "GroupAdded", "GroupRemoved", "GroupUpdated",
    "AddedToGroup", "RemovedFromGroup",
    # Site / admin ops surfaced under SharePoint workload
    "SiteCollectionAdminAdded", "SiteCollectionAdminRemoved",
    "SiteAdminAdded", "SiteAdminRemoved",
    # Extended file ops seen in real UAL exports
    "FileModifiedExtended", "FileAccessedExtended", "FileUploadedPartial",
    "FolderRenamed",
    # Sharing link usage — HIGH VALUE for data exfil detection
    "SharingLinkUsed", "SharingLinkCreated",
    # Permission changes
    "PermissionLevelAdded", "PermissionLevelRemoved", "PermissionLevelUpdated",
    "SharingInheritanceBroken", "SharingInheritanceReset",
    # List ops
    "ListColumnCreated", "ListColumnDeleted",
    "ListViewed", "ListItemViewed",
    # Site access
    "SiteAccessWithCustomScripts",
    # Search and page views (noisy but sometimes useful)
    "SearchQueryPerformed",
    "PageViewed", "PageViewedExtended",
}
# FIX: Added inbox rule variations seen in real UAL exports.
# "MoveToFolder" as a rule action appears under UpdateInboxRules JSON.
OP_INBOX_RULES = {
    "New-InboxRule", "Set-InboxRule", "Remove-InboxRule",
    "UpdateInboxRules",           # bulk/client-side rule update
    "Enable-InboxRule", "Disable-InboxRule",
    "Set-SweepRule", "Remove-SweepRule",  # Sweep rules (Outlook.com/OWA)
}
# FIX: Added Azure AD STS logon ops and common failure variants.
# parse_sign_in handles both Actor[] (AzureAD) and plain UserId (Exchange).
OP_SIGN_IN = {
    "UserLoggedIn", "UserLoginFailed",
    "PasswordLogonInitial", "PasswordLogonFailed",
    "UserLoggedOut",
    "SignInEvent",       # Azure AD / EOP sign-in variant
}
# FIX: Significantly expanded admin op set with BEC-relevant cmdlets.
# Many of these indicate persistence or privilege escalation.
OP_ADMIN = {
    # Role management
    "Add member to role.", "Remove member from role.",
    "Add delegate to role.", "Remove delegate from role.",
    "New-ManagementRoleAssignment", "Remove-ManagementRoleAssignment",
    # Mailbox configuration — BEC actors modify these to maintain access
    "Set-Mailbox", "Set-CASMailbox",
    "Set-MailboxAutoReplyConfiguration",  # auto-reply exfil
    "Set-MailboxMessageConfiguration",
    "Set-MailboxCalendarConfiguration",
    "Set-OwaMailboxPolicy",
    "Set-Clutter", "Set-FocusedInbox",   # sometimes used to suppress alerts
    # Mailbox permissions — classic BEC persistence
    "Add-MailboxPermission", "Remove-MailboxPermission",
    "Add-MailboxFolderPermission", "Remove-MailboxFolderPermission",
    "Update-MailboxFolderPermission",
    # App/OAuth — increasingly used for post-compromise persistence
    "New-ApplicationAccessPolicy", "Remove-ApplicationAccessPolicy",
    "Add service principal.", "Remove service principal.",
    "Add delegation entry.", "Remove delegation entry.",
    # Transport rules — redirect/copy flow
    "New-TransportRule", "Set-TransportRule", "Remove-TransportRule",
    "Enable-TransportRule", "Disable-TransportRule",
    # Audit/compliance — adversaries sometimes disable to reduce visibility
    "Set-AdminAuditLogConfig",
    "Set-OrganizationConfig",
    # SharePoint site administration
    "SiteCollectionCreated", "SiteCollectionDeleted",
    "SiteLocksChanged", "SiteCollectionQuotaModified",
    # EOP/hygiene admin cmdlets
    "Get-QuarantineMessage", "Release-QuarantineMessage",
    "New-QuarantinePolicy", "Set-QuarantinePolicy",
}
# FIX: Added Teams operations — a growing BEC vector for comms interception
# and social engineering within the org.
OP_TEAMS = {
    "MemberAdded", "MemberRemoved", "MemberRoleChanged",
    "TeamCreated", "TeamDeleted", "TeamUpdated",
    "ChannelAdded", "ChannelDeleted", "ChannelUpdated",
    "ChatCreated", "ChatRetrieved", "ChatUpdated",
    "MessageSent", "MessageDeleted", "MessageUpdated",
    "MessageHostedContentsListed", "MessageHostedContentRead",
    "BotAddedToTeam", "BotRemovedFromTeam",
    "AppInstalled", "AppUninstalled", "AppUpgraded",
    "ConnectorAdded", "ConnectorRemoved", "ConnectorUpdated",
    "TabAdded", "TabRemoved", "TabUpdated",
    "MeetingCreated", "MeetingDeleted",
    # Session and messaging variants seen in real exports
    "TeamsSessionStarted",
    "ReactedToMessage",
    "MessageCreatedHasLink", "MessageEditedHasLink",
    "ChatMessageCreated", "ChatMessageEdited", "ChatMessageDeleted",
}
# UPDATE ops that appear in multiple workloads — route by Workload in classifier
OP_MAILBOX_UPDATE = {
    "Update",   # Exchange: calendar/contact/task item updates
    "Move",     # Exchange: generic item move (not always has Subject)
}


def classify_operation(operation: str, workload: str = "") -> str:
    op = operation.strip()
    # Exact-match sets — ordered by BEC priority
    if op in OP_INBOX_RULES:   return "inbox_rule"
    if op in OP_MAIL_ACCESS:   return "mail_access"
    if op in OP_DELETE:        return "delete"
    if op in OP_EMAIL_SEND:    return "email_send"
    if op in OP_SIGN_IN:       return "sign_in"
    if op in OP_ADMIN:         return "admin"
    if op in OP_TEAMS:         return "teams"
    if op in OP_FILE_FOLDER:   return "file_folder"
    # FIX: "Update" and "Move" are ambiguous — use Workload to route correctly.
    # Exchange workload + Update/Move = mailbox item update (email_send parser).
    # Other workloads = generic.
    if op in OP_MAILBOX_UPDATE:
        wl = workload.strip().lower()
        if wl in ("exchange", "exchangeitem"):
            return "email_send"
        return "generic"
    return "generic"


def fmt_datetime(raw: str) -> str:
    """Convert any M365 UAL timestamp to dd MMM yyyy HH:MM:SS (24-hour).

    Handles all variants seen in real UAL exports:
      ISO 8601:  2026-03-18T22:38:40.0000000Z
                 2026-03-18T22:38:40Z
                 2026-03-18T22:38:40+11:00
                 2026-03-18 22:38:40
      AU slash:  16/04/2026 01:27        (no seconds — common in AU Excel exports)
                 16/04/2026 01:27:00
                 16/04/2026 1:27:00 AM
      US slash:  3/18/2026 10:38:40 PM  (Hawk / PowerShell exports)
                 3/18/2026 22:38:40
      Slash ISO: 2026/03/18 22:38:40
    Returns the original string unchanged if it cannot be parsed.
    """
    from datetime import datetime as _dt
    import re as _re

    if not raw or not isinstance(raw, str):
        return str(raw) if raw else ""
    s = raw.strip()
    if not s:
        return ""

    TARGET = "%d %b %Y %H:%M:%S"

    # ── Path 1: ISO 8601 — YYYY-MM-DD[T ]HH:MM:SS[.frac][Z/offset] ──────────
    if _re.match(r"\d{4}-\d{2}-\d{2}", s):
        s2 = _re.sub(r"\.\d+", "", s)              # drop .0000000 fractional
        s2 = _re.sub(r"Z$", "", s2)                  # drop trailing Z
        s2 = _re.sub(r"[+-]\d{2}:\d{2}$", "", s2) # drop tz offset
        s2 = s2.replace(" ", "T").replace("/", "-")  # normalise separators
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
            try:
                return _dt.strptime(s2, fmt).strftime(TARGET)
            except ValueError:
                continue

    # ── Path 2: Slash ISO — YYYY/MM/DD HH:MM:SS ──────────────────────────────
    if _re.match(r"\d{4}/\d{2}/\d{2}", s):
        s2 = s.replace("/", "-")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return _dt.strptime(s2, fmt).strftime(TARGET)
            except ValueError:
                continue

    # ── Path 3: D/M/YYYY or DD/MM/YYYY — AU locale (day-first) ──────────────
    # Must be checked before US path. If the day component > 12 it is
    # unambiguously DD/MM/YYYY. If <= 12 we still default to DD/MM/YYYY
    # because this tool is AU-focused and M365 AU tenants export day-first.
    if _re.match(r"\d{1,2}/\d{1,2}/\d{4}", s):
        for fmt in (
            "%d/%m/%Y %H:%M:%S",    # 16/04/2026 01:27:00
            "%d/%m/%Y %H:%M",       # 16/04/2026 01:27  (no seconds — real screenshot)
            "%d/%m/%Y %I:%M:%S %p", # 16/04/2026 1:27:00 AM
            "%d/%m/%Y %I:%M %p",    # 16/04/2026 1:27 AM
            "%d/%m/%Y",             # 16/04/2026
        ):
            try:
                return _dt.strptime(s, fmt).strftime(TARGET)
            except ValueError:
                continue

        # Fallback: try US month-first (for exports from non-AU tenants)
        for fmt in (
            "%m/%d/%Y %I:%M:%S %p", # 3/18/2026 10:38:40 PM
            "%m/%d/%Y %H:%M:%S",    # 3/18/2026 22:38:40
            "%m/%d/%Y %I:%M %p",    # 3/18/2026 10:38 PM
            "%m/%d/%Y %H:%M",       # 3/18/2026 22:38
            "%m/%d/%Y",             # 3/18/2026
        ):
            try:
                return _dt.strptime(s, fmt).strftime(TARGET)
            except ValueError:
                continue

    return raw  # unparseable — return original unchanged


def safe_get(d, *keys, default=""):
    """Safely traverse nested dict/list."""
    cur = d
    for k in keys:
        if cur is None:
            return default
        if isinstance(cur, dict):
            cur = cur.get(k)
        elif isinstance(cur, list) and isinstance(k, int):
            cur = cur[k] if k < len(cur) else None
        else:
            return default
    return cur if cur is not None else default


def join_list(lst, sep="; "):
    if not lst:
        return ""
    return sep.join(str(x) for x in lst if x)


# ─── Per-operation parsers ────────────────────────────────────────────────────

# AppAccessContext signature: present as a key, no Item / AffectedItems.
# Seen on Send, HardDelete, MailItemsAccessed when accessed by OAuth app token.
# These are NOT parse failures — distinct schema. Detect and return meaningful
# values so REQUIRED_FIELDS checks don't fire false WARNING floods.
_APP_ACCESS_KEYS = {"AppAccessContext", "ActorInfoString", "AuthType"}

def _is_app_access_context(audit: dict) -> bool:
    """True when AuditData is an OAuth app-access record with no mail content."""
    has_marker = bool(_APP_ACCESS_KEYS & set(audit.keys()))
    has_content = "Item" in audit or "AffectedItems" in audit
    return has_marker and not has_content

def _app_access_extra(audit: dict) -> str:
    """Extract what we CAN from an AppAccessContext record."""
    parts = []
    ctx = audit.get("AppAccessContext", {}) or {}
    if isinstance(ctx, dict):
        aac_id = ctx.get("AADSessionId", "") or ctx.get("UniqueTokenId", "")
        if aac_id:
            parts.append(f"AADSessionId={aac_id}")
    app_id = audit.get("AppId", "")
    if app_id:
        parts.append(f"AppId={app_id}")
    actor = audit.get("ActorInfoString", "")
    if actor:
        parts.append(f"Actor={actor[:80]}")
    auth  = audit.get("AuthType", "")
    if auth:
        parts.append(f"AuthType={auth}")
    return " | ".join(parts) if parts else "AppAccessContext record — no mail content available"


def parse_email_send(audit: dict) -> dict:
    if _is_app_access_context(audit):
        return {
            "Subject":          "(App/OAuth access - no item content in AuditData)",
            "Path/Folder":      "",
            "Size (bytes)":     "",
            "InternetMessageId":"",
            "Attachments":      "",
            "Extra":            _app_access_extra(audit),
        }
    item = audit.get("Item", {}) or {}
    parent = item.get("ParentFolder", {}) or {}
    attachments = item.get("Attachments", []) or []
    att_names = join_list([a.get("Name", "") for a in attachments if isinstance(a, dict)])
    return {
        "Subject":          item.get("Subject", "(No Subject)"),
        "Path/Folder":      parent.get("Path", "(No Path)"),
        "Size (bytes)":     item.get("Size", ""),
        "InternetMessageId":item.get("InternetMessageId", ""),
        "Attachments":      att_names,
        "Extra":            "",
    }


def parse_delete(audit: dict) -> dict:
    if _is_app_access_context(audit):
        return {
            "Subject":          "(App/OAuth access - no item content in AuditData)",
            "Path/Folder":      "(App/OAuth access - no item content in AuditData)",
            "Size (bytes)":     "",
            "InternetMessageId":"(App/OAuth access - no item content in AuditData)",
            "Attachments":      "",
            "Extra":            _app_access_extra(audit),
        }
    affected = audit.get("AffectedItems", []) or []
    paths, subjects, msg_ids = [], [], []
    for a in affected:
        if not isinstance(a, dict):
            continue
        pf = a.get("ParentFolder", {}) or {}
        paths.append(pf.get("Path", ""))
        subjects.append(a.get("Subject", ""))
        msg_ids.append(a.get("InternetMessageId", ""))
    return {
        "Subject":           join_list(subjects),
        "Path/Folder":       join_list(paths),
        "Size (bytes)":      "",
        "InternetMessageId": join_list(msg_ids),
        "Attachments":       "",
        "Extra":             f"{len(affected)} item(s) affected",
    }


def parse_mail_access(audit: dict) -> dict:
    if _is_app_access_context(audit):
        return {
            "Subject":          "(App/OAuth access - no item content in AuditData)",
            "Path/Folder":      "(App/OAuth access - no item content in AuditData)",
            "Size (bytes)":     "",
            "InternetMessageId":"(App/OAuth access - no item content in AuditData)",
            "Attachments":      "",
            "Extra":            _app_access_extra(audit),
        }
    affected = audit.get("AffectedItems", []) or []
    paths, msg_ids = [], []
    for a in affected:
        if not isinstance(a, dict):
            continue
        pf = a.get("ParentFolder", {}) or {}
        paths.append(pf.get("Path", ""))
        msg_ids.append(a.get("InternetMessageId", ""))
    access_type  = audit.get("MailAccessType", "")
    client       = audit.get("AppId", "") or audit.get("ClientAppId", "")
    protocols    = audit.get("ClientInfoString", "")
    external     = audit.get("ExternalAccess", "")
    # Aggregate/sync record: ExternalAccess present but no AffectedItems
    # This is a batched access — InternetMessageId intentionally absent
    is_aggregate = (not affected and external != "")
    if is_aggregate:
        subject_note = "(Aggregate sync record - no individual message IDs available)"
    else:
        subject_note = "(See InternetMessageId – correlate via Message Trace)"
    extra_parts = [
        f"AccessType={access_type}" if access_type else "",
        f"AppId={client}"           if client      else "",
        f"Protocol={protocols}"     if protocols   else "",
        f"ExternalAccess={external}" if external != "" else "",
        "AGGREGATE_RECORD"          if is_aggregate else "",
    ]
    return {
        "Subject":           subject_note,
        "Path/Folder":       join_list(paths),
        "Size (bytes)":      "",
        "InternetMessageId": join_list(msg_ids),
        "Attachments":       "",
        "Extra":             join_list([p for p in extra_parts if p], " | "),
    }


def parse_file_folder(audit: dict) -> dict:
    filename   = audit.get("SourceFileName", audit.get("DestinationFileName", "(No FileName)"))
    obj_id     = audit.get("ObjectId", "(No ObjectId)")
    site       = audit.get("SiteUrl", "")
    src_rel    = audit.get("SourceRelativeUrl", "")
    dst_rel    = audit.get("DestinationRelativeUrl", "")
    return {
        "Subject":           filename,
        "Path/Folder":       obj_id,
        "Size (bytes)":      audit.get("FileSize", ""),
        "InternetMessageId": "",
        "Attachments":       "",
        "Extra":             join_list([site, src_rel, dst_rel], " | "),
    }


def parse_inbox_rule(audit: dict) -> dict:
    rule_name = "(No Rule Name)"
    params = audit.get("Parameters", []) or []
    for p in params:
        if isinstance(p, dict) and p.get("Name") == "Name":
            rule_name = p.get("Value", rule_name)
            break
    obj_id = audit.get("ObjectId", "(No ObjectId)")
    # Also try to grab rule conditions for context
    conditions = []
    for p in params:
        if isinstance(p, dict) and p.get("Name") in (
            "MoveToFolder", "SubjectContainsWords", "FromAddressContainsWords",
            "DeleteMessage", "ForwardTo", "RedirectTo", "MarkAsRead"
        ):
            conditions.append(f"{p['Name']}={p.get('Value','')}")
    return {
        "Subject":           rule_name,
        "Path/Folder":       obj_id,
        "Size (bytes)":      "",
        "InternetMessageId": "",
        "Attachments":       "",
        "Extra":             join_list(conditions, " | ") if conditions else "",
    }


def parse_sign_in(audit: dict) -> dict:
    # FIX: Actor[] array is only present in AzureActiveDirectoryStsLogon records.
    # Exchange-sourced logon records have UserId at the top level only.
    # Try Actor[] first (Type 5 = UPN), then fall through to top-level UserId.
    actor_list = audit.get("Actor", []) or []
    actor = next((a.get("ID", "") for a in actor_list if isinstance(a, dict) and a.get("Type") == 5), "")
    if not actor:
        actor = audit.get("UserId", "")
    error_num  = audit.get("ErrorNumber", "")
    log_on_err = audit.get("LogonError", "")
    client_ip  = audit.get("ClientIP", audit.get("ActorIpAddress", ""))
    user_agent = audit.get("DeviceProperties", [])
    ua_str = ""
    if isinstance(user_agent, list):
        ua_str = next((p.get("Value","") for p in user_agent if isinstance(p,dict) and p.get("Name") == "UserAgent"), "")
    return {
        "Subject":           actor or "(Unknown Actor)",
        "Path/Folder":       audit.get("OrganizationId", ""),
        "Size (bytes)":      "",
        "InternetMessageId": "",
        "Attachments":       "",
        "Extra":             f"ErrorNumber={error_num} | LogonError={log_on_err} | ClientIP={client_ip} | UA={ua_str}",
    }


def parse_admin(audit: dict) -> dict:
    params = audit.get("Parameters", []) or []
    param_str = join_list([f"{p.get('Name')}={p.get('Value','')}" for p in params if isinstance(p, dict)], " | ")
    obj_id = audit.get("ObjectId", "")
    return {
        "Subject":           obj_id or "(No ObjectId)",
        "Path/Folder":       audit.get("OrganizationName", ""),
        "Size (bytes)":      "",
        "InternetMessageId": "",
        "Attachments":       "",
        "Extra":             param_str,
    }


def parse_teams(audit: dict) -> dict:
    """Teams ops: extract channel, chat, membership, and message context."""
    team_name    = audit.get("TeamName", "")
    channel_name = audit.get("ChannelName", "")
    chat_id      = audit.get("CommunicationType", "")
    members      = audit.get("Members", []) or []
    member_upns  = join_list([m.get("UPN", m.get("DisplayName", "")) for m in members if isinstance(m, dict)])
    msg_content  = audit.get("MessageContent", "")  # sometimes populated in premium audit
    app_name     = audit.get("AddonName", audit.get("Name", ""))
    return {
        "Subject":           team_name or channel_name or chat_id or app_name or "(Teams Op)",
        "Path/Folder":       channel_name or "",
        "Size (bytes)":      "",
        "InternetMessageId": "",
        "Attachments":       app_name if app_name else "",
        "Extra":             f"Members={member_upns}" if member_upns else msg_content[:200] if msg_content else "",
    }


def parse_generic(audit: dict) -> dict:
    return {
        "Subject":           str(audit.get("ObjectId", "(Unknown)")),
        "Path/Folder":       "",
        "Size (bytes)":      "",
        "InternetMessageId": "",
        "Attachments":       "",
        "Extra":             "",
    }


PARSERS = {
    "email_send": parse_email_send,
    "delete":     parse_delete,
    "mail_access": parse_mail_access,
    "file_folder": parse_file_folder,
    "inbox_rule":  parse_inbox_rule,
    "sign_in":     parse_sign_in,
    "teams":       parse_teams,
    "admin":       parse_admin,
    "generic":     parse_generic,
}


# ─── Core parse function ──────────────────────────────────────────────────────

OUTPUT_COLUMNS = [
    "Row", "CreationTime", "Id", "Operation", "OperationCategory",
    "UserId", "ClientIP", "Workload", "ResultStatus",
    "Subject", "Path/Folder", "Size (bytes)", "InternetMessageId",
    "Attachments", "Extra", "RawAuditData"
]


# ─── Error tracking ──────────────────────────────────────────────────────────

# Error severity levels used in ParseError.severity
ERR_CRITICAL = "CRITICAL"   # Row completely unparseable — analyst must review raw data
ERR_WARN     = "WARNING"    # Row parsed but key fields are empty/placeholder
ERR_INFO     = "INFO"       # Non-fatal anomaly — logged but row likely usable

# Placeholder strings that indicate a silent parse failure
# Placeholders that indicate a genuine parse failure / missing data → trigger WARNING
WARN_PLACEHOLDERS = {
    "(No Subject)", "(No Path)", "(No Item)", "(No Folder)",
    "(No SourceFileName)", "(No ObjectId)", "(No Rule Name)", "(No Item Key)",
    "(No AffectedItems)", "(Unknown)",
}
# Placeholders that are intentional informational notes → suppress WARNING
# These mean "we parsed correctly but this schema variant has no extractable content"
INFO_PLACEHOLDERS = {
    "(Unknown Actor)",
    "(See InternetMessageId – correlate via Message Trace)",
    "(Aggregate sync record - no individual message IDs available)",
    "(App/OAuth access - no item content in AuditData)",
}
# Combined set used for CRITICAL row annotation check
EMPTY_PLACEHOLDERS = WARN_PLACEHOLDERS | INFO_PLACEHOLDERS


class ParseError:
    """Structured record of a parse failure or anomaly for a single UAL row."""
    __slots__ = ("row_num", "operation", "user_id", "creation_time",
                 "severity", "error_type", "detail", "raw_audit_data")

    def __init__(self, row_num, operation, user_id, creation_time,
                 severity, error_type, detail, raw_audit_data=""):
        self.row_num       = row_num
        self.operation     = operation
        self.user_id       = user_id
        self.creation_time = creation_time
        self.severity      = severity
        self.error_type    = error_type
        self.detail        = detail
        self.raw_audit_data = raw_audit_data

    def to_dict(self) -> dict:
        return {
            "Row":           self.row_num,
            "Severity":      self.severity,
            "ErrorType":     self.error_type,
            "Detail":        self.detail,
            "Operation":     self.operation,
            "UserId":        self.user_id,
            "CreationTime":  self.creation_time,
            "RawAuditData":  self.raw_audit_data,
        }


ERROR_COLUMNS = [
    "Row", "Severity", "ErrorType", "Detail",
    "Operation", "UserId", "CreationTime", "RawAuditData"
]

# Silent-failure detectors: fields that should NOT be empty/placeholder
# for each operation category. Analyst-facing — if these are blank after
# parsing, it means the AuditData schema didn't match expectations.
REQUIRED_FIELDS_BY_CATEGORY = {
    "email_send":  ["Subject"],
    "delete":      ["Path/Folder"],
    "mail_access": ["InternetMessageId"],
    "file_folder": ["Subject"],          # SourceFileName
    "inbox_rule":  ["Subject"],          # Rule name
    "sign_in":     [],                   # UPN best-effort; absent on system/synthetic sign-in events
    "admin":       [],                   # ObjectId often absent on EOP/hygiene ops
    "teams":       [],                   # Schema varies too much to enforce
    "generic":     [],
}


def _normalise_record(raw: dict) -> dict:
    """Case-insensitive column normalisation for varied UAL export formats.
    Handles: AuditData / auditdata / Audit Data
             Operation / Operations (PowerShell plural form)
             UserId / UserIds / User Id
             CreationTime / CreationDate
    """
    mapping = {}
    for k, v in raw.items():
        lk = k.strip().lower().replace(" ", "").replace("_", "")
        mapping[lk] = v
    result = {}
    for canonical, variants in [
        ("AuditData",    ["auditdata"]),
        ("Operation",    ["operation", "operations"]),
        ("CreationTime", ["creationtime", "creationdate"]),
        ("UserId",       ["userid", "userids", "user"]),
        ("ClientIP",     ["clientip", "clientipaddress"]),
        ("Workload",     ["workload"]),
        ("ResultStatus", ["resultstatus", "result"]),
        ("Id",           ["id", "auditrecordid"]),
    ]:
        for v in variants:
            if v in mapping:
                result[canonical] = mapping[v]
                break
        else:
            result[canonical] = ""
    return result


def parse_row(row_num: int, raw_record: dict) -> tuple:
    """Parse a single UAL record into (output_dict, list[ParseError]).

    Always returns a row dict — even on failure — so the analyst has
    every row in the output. Errors are accumulated separately and
    written to a dedicated error sheet / CSV for review.

    Error hierarchy:
      CRITICAL — AuditData could not be JSON-decoded at all. The row
                 is written with [JSON PARSE FAILURE] markers so it is
                 immediately visible in the output. Analyst must check
                 the RawAuditData column manually.
      WARNING  — AuditData decoded but a required field for this
                 operation type came back empty or as a known placeholder.
                 Often means the AuditData schema variant is not yet
                 handled, or the record is genuinely missing data.
      INFO     — Structural anomaly that doesn't affect usability:
                 e.g. operation not in any known set (new/unknown op),
                 or AuditData was an empty object {}.
    """
    errors: list = []
    norm = _normalise_record(raw_record)
    operation  = norm.get("Operation", "") or raw_record.get("Operation", "")
    workload   = norm.get("Workload", "")  or raw_record.get("Workload", "")
    audit_data = norm.get("AuditData", "") or raw_record.get("AuditData", "{}")
    if not audit_data:
        audit_data = "{}"

    raw_str = audit_data if not isinstance(audit_data, dict) else json.dumps(audit_data)
    creation_time = norm.get("CreationTime") or ""
    user_id = norm.get("UserId") or ""

    # If the top-level CreationTime has no seconds component (e.g. "16/04/2026 01:27"
    # from an Excel-resaved CSV), check whether AuditData JSON has a higher-precision
    # timestamp and prefer it. Purview's AuditData.CreationTime is always full ISO 8601.
    _has_seconds = ":" in creation_time[creation_time.rfind(":")-3:] if creation_time and creation_time.count(":") >= 2 else False
    if creation_time and not _has_seconds:
        try:
            _ad_quick = json.loads(audit_data) if not isinstance(audit_data, dict) else audit_data
            _audit_ct = _ad_quick.get("CreationTime", "") if isinstance(_ad_quick, dict) else ""
            if _audit_ct and ":" in _audit_ct[_audit_ct.rfind(":")-3:]:
                creation_time = _audit_ct  # use full-precision AuditData timestamp
        except Exception:
            pass

    # FIX: Workload is often blank in the CSV wrapper — decode AuditData early
    # just to pull Workload from it before classifying. Full decode happens below.
    if not workload and audit_data and audit_data != "{}":
        try:
            _quick = json.loads(audit_data) if not isinstance(audit_data, dict) else audit_data
            workload = _quick.get("Workload", "") if isinstance(_quick, dict) else ""
        except Exception:
            pass

    category   = classify_operation(operation, workload)

    def _err(severity, error_type, detail):
        errors.append(ParseError(
            row_num=row_num, operation=operation, user_id=user_id,
            creation_time=creation_time, severity=severity,
            error_type=error_type, detail=detail, raw_audit_data=raw_str
        ))

    # ── Step 1: Decode AuditData JSON ────────────────────────────────────────
    json_parse_failed = False
    if isinstance(audit_data, dict):
        audit = audit_data
    else:
        try:
            audit = json.loads(audit_data)
            if not isinstance(audit, dict):
                # Valid JSON but not an object (e.g. a bare array or string)
                _err(ERR_CRITICAL, "INVALID_AUDIT_SCHEMA",
                     f"AuditData decoded to {type(audit).__name__}, expected dict. "
                     f"Value preview: {str(audit)[:120]}")
                audit = {}
                json_parse_failed = True
        except json.JSONDecodeError as e:
            _err(ERR_CRITICAL, "JSON_PARSE_FAILURE",
                 f"AuditData could not be decoded: {e}. "
                 f"Raw preview: {str(audit_data)[:200]}")
            audit = {}
            json_parse_failed = True
        except Exception as e:
            _err(ERR_CRITICAL, "AUDITDATA_DECODE_ERROR",
                 f"Unexpected error decoding AuditData: {type(e).__name__}: {e}")
            audit = {}
            json_parse_failed = True

    # ── Step 2: Warn if AuditData is empty (decoded fine but has no content) ─
    if not json_parse_failed and audit == {}:
        _err(ERR_INFO, "EMPTY_AUDITDATA",
             "AuditData field is present but decodes to an empty object {}. "
             "This can occur for synthetic/system-generated UAL records.")

    # ── Step 3: Warn on unknown operation (falls through to generic) ──────────
    if category == "generic" and operation and operation not in ("", "generic"):
        _err(ERR_INFO, "UNKNOWN_OPERATION",
             f"Operation '{operation}' (Workload: '{workload}') is not in any "
             f"known category set. Routed to generic parser. If this is a "
             f"BEC-relevant operation, add it to the appropriate OP_ set.")

    # ── Step 4: Run the category parser with per-field exception isolation ────
    try:
        parsed = PARSERS.get(category, parse_generic)(audit)
    except Exception as e:
        _err(ERR_CRITICAL, "PARSER_EXCEPTION",
             f"Parser for category '{category}' raised {type(e).__name__}: {e}. "
             f"AuditData keys: {list(audit.keys())[:20]}")
        parsed = {
            "Subject": "[PARSER EXCEPTION — SEE ERROR LOG]",
            "Path/Folder": "", "Size (bytes)": "",
            "InternetMessageId": "", "Attachments": "", "Extra": str(e),
        }

    # ── Step 5: Check for silent failures (placeholder / empty required fields) ─
    if not json_parse_failed:
        for field in REQUIRED_FIELDS_BY_CATEGORY.get(category, []):
            val = str(parsed.get(field, "")).strip()
            if not val or val in WARN_PLACEHOLDERS:
                _err(ERR_WARN, "EMPTY_REQUIRED_FIELD",
                     f"Field '{field}' is empty or placeholder ('{val}') for "
                     f"category '{category}'. AuditData keys present: "
                     f"{list(audit.keys())[:15]}. This may indicate an "
                     f"unhandled AuditData schema variant.")

    # ── Step 6: Annotate output row if CRITICAL errors exist ─────────────────
    critical_flag = ""
    if any(e.severity == ERR_CRITICAL for e in errors):
        critical_flag = "[JSON PARSE FAILURE — REVIEW RAW]"
        if not parsed.get("Subject") or parsed["Subject"] in EMPTY_PLACEHOLDERS:
            parsed["Subject"] = critical_flag

    out_row = {
        "Row":              row_num,
        "CreationTime":     fmt_datetime(creation_time or audit.get("CreationTime", "")),
        "Id":               norm.get("Id") or audit.get("Id", ""),
        "Operation":        operation,
        "OperationCategory": category,
        "UserId":           user_id or audit.get("UserId", ""),
        "ClientIP":         norm.get("ClientIP") or audit.get("ClientIPAddress", audit.get("ClientIP", "")),
        "Workload":         workload or audit.get("Workload", ""),
        "ResultStatus":     norm.get("ResultStatus") or audit.get("ResultStatus", ""),
        "Subject":          parsed["Subject"],
        "Path/Folder":      parsed["Path/Folder"],
        "Size (bytes)":     parsed["Size (bytes)"],
        "InternetMessageId":parsed["InternetMessageId"],
        "Attachments":      parsed["Attachments"],
        "Extra":            parsed["Extra"],
        "RawAuditData":     raw_str,
        "_parse_errors":    errors,   # stripped before write, used for error sheet
    }
    return out_row, errors


def load_input_file(filepath: str) -> list[dict]:
    """Load CSV or XLSX UAL export. Returns list of row dicts."""
    ext = Path(filepath).suffix.lower()
    rows = []

    if ext == ".csv":
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

    elif ext in (".xlsx", ".xls") and HAS_OPENPYXL:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else f"Col{i}" for i, c in enumerate(row)]
            else:
                rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))
        wb.close()

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return rows


def write_error_xlsx_sheet(wb, error_list: list):
    """Add a plain B&W Errors sheet with severity-coded left border accent."""
    ws = wb.create_sheet("⚠ Parse Errors")

    hdr_fill  = PatternFill("solid", fgColor="D9D9D9")
    crit_fill = PatternFill("solid", fgColor="FFE0E0")   # light red tint
    warn_fill = PatternFill("solid", fgColor="FFF3CD")   # light amber tint
    info_fill = PatternFill("solid", fgColor="FFFFFF")   # plain white
    hdr_font  = Font(bold=True, color="000000", size=10, name="Calibri")
    cell_font = Font(color="000000", size=9, name="Calibri")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, bottom=thin, top=thin)

    ws.append(ERROR_COLUMNS)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 24

    sev_fill = {ERR_CRITICAL: crit_fill, ERR_WARN: warn_fill, ERR_INFO: info_fill}
    for err in error_list:
        d = err.to_dict()
        ws.append([str(d.get(c, "")) for c in ERROR_COLUMNS])
        row_idx = ws.max_row
        fill = sev_fill.get(err.severity, info_fill)
        for col in range(1, len(ERROR_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = cell_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        ws.row_dimensions[row_idx].height = 18

    col_widths = {
        "Row": 6, "Severity": 12, "ErrorType": 28, "Detail": 70,
        "Operation": 22, "UserId": 30, "CreationTime": 24, "RawAuditData": 22
    }
    for col, name in enumerate(ERROR_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(name, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def write_output_xlsx(output_path: str, parsed_rows: list[dict]):
    """Write parsed rows to a styled XLSX."""
    from openpyxl.utils.datetime import to_excel as _to_excel
    import re as _re2
    from datetime import datetime as _dt2

    def _parse_ct_for_serial(raw):
        """Parse any UAL-formatted timestamp to datetime for serial conversion."""
        if not raw or not isinstance(raw, str):
            return None
        s = raw.strip()
        for fmt in (
            "%d %b %Y %H:%M:%S",          # already-formatted: 18 Mar 2026 22:38:40
            "%d %b %Y %H:%M",
        ):
            try: return _dt2.strptime(s, fmt)
            except ValueError: continue
        return None

    # Strip internal error tracking field before writing
    clean_rows = [{k: v for k, v in r.items() if k != "_parse_errors"} for r in parsed_rows]
    all_errors = [e for r in parsed_rows for e in r.get("_parse_errors", [])]

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "UAL Parsed"

    # ── Plain black-and-white styling ───────────────────────────────────────
    hdr_font  = Font(bold=True, color="000000", size=10, name="Calibri")
    cell_font = Font(color="000000", size=9, name="Calibri")
    hdr_fill  = PatternFill("solid", fgColor="D9D9D9")   # light grey header
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, bottom=thin, top=thin)
    even_fill = PatternFill("solid", fgColor="F2F2F2")   # very light grey alternating rows
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")   # white

    # Write header
    ws.append(OUTPUT_COLUMNS)
    for col, cell in enumerate(ws[1], 1):
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    ws.row_dimensions[1].height = 30

    # Write data rows — alternating white / light grey, no colour coding
    # Pre-compute CreationTime column index once
    _ct_col = OUTPUT_COLUMNS.index("CreationTime") + 1


    for idx, row in enumerate(clean_rows, 2):
        fill = even_fill if idx % 2 == 0 else odd_fill
        for col_name in OUTPUT_COLUMNS:
            col_idx = OUTPUT_COLUMNS.index(col_name) + 1
            raw_val = row.get(col_name, "")

            if col_name == "CreationTime" and raw_val:
                parsed_dt = _parse_ct_for_serial(str(raw_val))
                if parsed_dt:
                    cell = ws.cell(row=idx, column=col_idx,
                                   value=_to_excel(parsed_dt))
                    # font MUST be set before number_format.
                    # Setting font initialises cell._style and the parent chain
                    # (cell -> ws -> wb) that number_format descriptor requires
                    # to register the custom format ID correctly in the XML.
                    # Setting number_format first leaves numFmtId=0 (General).
                    cell.font          = cell_font
                    cell.number_format = "DD MMM YYYY HH:MM:SS"
                else:
                    cell = ws.cell(row=idx, column=col_idx, value=str(raw_val))
                    cell.font = cell_font
            else:
                cell = ws.cell(row=idx, column=col_idx, value=str(raw_val))
                cell.font = cell_font

            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        ws.row_dimensions[idx].height = 18


    # Column widths
    col_widths = {
        "Row": 5, "CreationTime": 24, "Id": 14, "Operation": 22,
        "OperationCategory": 16, "UserId": 28, "ClientIP": 18,
        "Workload": 14, "ResultStatus": 14, "Subject": 40,
        "Path/Folder": 38, "Size (bytes)": 12, "InternetMessageId": 36,
        "Attachments": 28, "Extra": 36, "RawAuditData": 18
    }
    for col, name in enumerate(OUTPUT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(name, 16)

    # Freeze pane
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Category", "Count"])
    from collections import Counter
    cats = Counter(r.get("OperationCategory", "generic") for r in clean_rows)
    for cat, count in sorted(cats.items()):
        ws2.append([cat, count])
    ws2.append(["TOTAL", len(parsed_rows)])

    # Add error counts to summary
    err_counts = Counter(e.severity for e in all_errors)
    ws2.append([])
    ws2.append(["--- Parse Quality ---", ""])
    ws2.append(["CRITICAL errors", err_counts.get(ERR_CRITICAL, 0)])
    ws2.append(["WARNING errors",  err_counts.get(ERR_WARN, 0)])
    ws2.append(["INFO notices",    err_counts.get(ERR_INFO, 0)])
    ws2.append(["Rows with errors", len([r for r in parsed_rows if r.get("_parse_errors")])])

    # Add error sheet if any errors exist
    if all_errors:
        write_error_xlsx_sheet(wb, all_errors)

    wb.save(output_path)
    return output_path, len(all_errors)


def write_output_csv(output_path: str, parsed_rows: list[dict]):
    clean_rows = [{k: v for k, v in r.items() if k != "_parse_errors"} for r in parsed_rows]
    all_errors = [e for r in parsed_rows for e in r.get("_parse_errors", [])]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(clean_rows)

    # Write companion error CSV alongside main output
    if all_errors:
        err_path = output_path.replace(".csv", "_ERRORS.csv")
        with open(err_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=ERROR_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows([e.to_dict() for e in all_errors])

    return output_path, len(all_errors)



def generate_html_report(parsed_rows: list, output_path: str, source_filename: str = "") -> str:
    """Generate a self-contained HTML investigation report from parsed UAL rows.

    Returns the path of the written HTML file.
    """
    import html as _html
    from collections import Counter
    from datetime import datetime as _dt

    esc = _html.escape

    total      = len(parsed_rows)
    cats       = Counter(r.get("OperationCategory", "generic") for r in parsed_rows)
    ops        = Counter(r.get("Operation", "")               for r in parsed_rows)
    users      = Counter(r.get("UserId", "")                  for r in parsed_rows)
    client_ips = Counter(r.get("ClientIP", "")                for r in parsed_rows)
    results    = Counter(r.get("ResultStatus", "")            for r in parsed_rows)
    generated  = _dt.now().strftime("%d %b %Y %H:%M:%S")

    # BEC severity map — matches ThreadHunter
    BEC_SEV = {
        "New-InboxRule": "CRITICAL", "UpdateInboxRules": "CRITICAL",
        "Set-InboxRule": "CRITICAL", "Remove-InboxRule": "CRITICAL",
        "Set-Mailbox": "CRITICAL", "Add-MailboxPermission": "CRITICAL",
        "SendAs": "HIGH", "Send": "HIGH",
        "FileDownloaded": "HIGH", "FileSyncDownloadedFull": "HIGH",
        "HardDelete": "HIGH", "SoftDelete": "HIGH",
        "MailItemsAccessed": "MEDIUM",
        "UserLoggedIn": "LOW", "UserLoginFailed": "LOW",
    }
    SEV_COLOR = {
        "CRITICAL": ("#FF4D6D", "#2d0010"),
        "HIGH":     ("#DCA032", "#2d1a00"),
        "MEDIUM":   ("#6374F8", "#0a0e2a"),
        "LOW":      ("#8C90A4", "#181b25"),
    }

    def sev_badge(op):
        sev = BEC_SEV.get(op, "")
        if sev and sev in SEV_COLOR:
            fg, bg = SEV_COLOR[sev]
            return (f'<span style="background:{bg};color:{fg};padding:2px 8px;' 
                    f'border-radius:3px;font-size:11px;font-weight:bold;' 
                    f'border:1px solid {fg};font-family:Courier New,monospace">{sev}</span>')
        return ""

    def stat_card(label, value, color="#6374F8", sub=""):
        return (f'<div class="stat-card">' 
                f'<div class="stat-val" style="color:{color}">{esc(str(value))}</div>' 
                f'<div class="stat-label">{esc(label)}</div>' 
                f'{f'<div class="stat-sub">{esc(sub)}</div>' if sub else ""}' 
                f'</div>')

    def bar_row(label, count, max_count, color="#6374F8", badge=""):
        pct = int((count / max(max_count, 1)) * 100)
        return (f'<tr><td class="bar-label">{esc(str(label))}</td>' 
                f'<td class="bar-cell">' 
                f'<div class="bar-track"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div>' 
                f'</td><td class="bar-count">{count:,}</td>' 
                f'<td class="bar-badge">{badge}</td></tr>')

    # Category colours matching the GUI
    CAT_COLOR = {
        "email_send":  "#2EAD7A",
        "delete":      "#E95555",
        "mail_access": "#A78BFA",
        "file_folder": "#60A5FA",
        "inbox_rule":  "#DCA032",
        "teams":       "#38BDF8",
        "sign_in":     "#34D399",
        "admin":       "#F87171",
        "generic":     "#8C90A4",
    }

    # ── Stat cards ────────────────────────────────────────────────────────────
    del_count  = cats.get("delete", 0)
    rule_count = cats.get("inbox_rule", 0)
    stats_html = (
        stat_card("Total Rows",     f"{total:,}",                          "#6374F8") +
        stat_card("Email Ops",      f"{cats.get('email_send', 0):,}",     "#2EAD7A") +
        stat_card("Mail Access",    f"{cats.get('mail_access', 0):,}",    "#A78BFA") +
        stat_card("Deletes",        f"{del_count:,}",
                  "#E95555" if del_count  > 0 else "#8C90A4") +
        stat_card("Inbox Rules",    f"{rule_count:,}",
                  "#DCA032" if rule_count > 0 else "#8C90A4") +
        stat_card("File/SharePoint",f"{cats.get('file_folder', 0):,}",    "#60A5FA") +
        stat_card("Teams",          f"{cats.get('teams', 0):,}",          "#38BDF8") +
        stat_card("Admin Ops",      f"{cats.get('admin', 0):,}",          "#F87171")
    )

    # ── Operation breakdown ───────────────────────────────────────────────────
    top_ops = ops.most_common(20)
    max_op  = top_ops[0][1] if top_ops else 1
    ops_rows = "".join(
        bar_row(op, cnt, max_op,
                CAT_COLOR.get(BEC_SEV.get(op, ""), "#6374F8")
                if op in BEC_SEV else "#6374F8",
                sev_badge(op))
        for op, cnt in top_ops if op
    )

    # ── Top users ─────────────────────────────────────────────────────────────
    top_users = [(u, c) for u, c in users.most_common(15) if u]
    max_usr   = top_users[0][1] if top_users else 1
    usr_rows  = "".join(bar_row(u, c, max_usr, "#2EAD7A") for u, c in top_users)

    # ── Top IPs ───────────────────────────────────────────────────────────────
    top_ips  = [(ip, c) for ip, c in client_ips.most_common(15) if ip]
    max_ip   = top_ips[0][1] if top_ips else 1
    ip_rows  = "".join(bar_row(ip, c, max_ip, "#E95555") for ip, c in top_ips)

    # ── Result status breakdown ───────────────────────────────────────────────
    top_res = [(r, c) for r, c in results.most_common(10) if r]
    max_res = top_res[0][1] if top_res else 1
    res_rows = "".join(
        bar_row(r, c, max_res,
                "#2EAD7A" if "succ" in r.lower() else "#E95555")
        for r, c in top_res
    )

    # ── Category breakdown ────────────────────────────────────────────────────
    cat_rows = "".join(
        bar_row(cat, cnt, total, CAT_COLOR.get(cat, "#6374F8"))
        for cat, cnt in sorted(cats.items(), key=lambda x: x[1], reverse=True)
    )

    # ── Inbox rule detail table ───────────────────────────────────────────────
    rule_rows_data = [r for r in parsed_rows if r.get("OperationCategory") == "inbox_rule"]
    if rule_rows_data:
        rule_tbl_rows = "".join(
            f'<tr><td>{esc(r.get("CreationTime",""))}</td>' 
            f'<td>{esc(r.get("UserId",""))}</td>' 
            f'<td>{esc(r.get("Operation",""))}</td>' 
            f'<td class="subj">{esc(r.get("Subject",""))}</td>' 
            f'<td class="subj">{esc(r.get("Extra","")[:120])}</td></tr>'
            for r in rule_rows_data[:50]
        )
        rule_section = f"""
        <section>
          <h2>Inbox Rule Operations <span class="badge-crit">⚠ {len(rule_rows_data)} event(s)</span></h2>
          <p class="note">Inbox rules are the primary BEC persistence mechanism — review all entries below.</p>
          <table>
            <thead><tr>
              <th>Timestamp</th><th>User</th><th>Operation</th>
              <th>Rule Name</th><th>Conditions / Actions</th>
            </tr></thead>
            <tbody>{rule_tbl_rows}</tbody>
          </table>
          {'<p class="note">Showing first 50 rows.</p>' if len(rule_rows_data) > 50 else ""}
        </section>"""
    else:
        rule_section = ""

    # ── Delete detail table ───────────────────────────────────────────────────
    del_rows_data = [r for r in parsed_rows if r.get("OperationCategory") == "delete"]
    if del_rows_data:
        del_tbl_rows = "".join(
            f'<tr><td>{esc(r.get("CreationTime",""))}</td>' 
            f'<td>{esc(r.get("UserId",""))}</td>' 
            f'<td>{esc(r.get("Operation",""))}</td>' 
            f'<td class="subj">{esc(r.get("Subject","")[:80])}</td>' 
            f'<td>{esc(r.get("Path/Folder","")[:80])}</td></tr>'
            for r in del_rows_data[:100]
        )
        del_section = f"""
        <section>
          <h2>Delete Operations <span class="badge-high">⚠ {len(del_rows_data)} event(s)</span></h2>
          <table>
            <thead><tr>
              <th>Timestamp</th><th>User</th><th>Operation</th>
              <th>Subject</th><th>Folder Path</th>
            </tr></thead>
            <tbody>{del_tbl_rows}</tbody>
          </table>
          {'<p class="note">Showing first 100 rows.</p>' if len(del_rows_data) > 100 else ""}
        </section>"""
    else:
        del_section = ""

    # ── Full HTML ─────────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>UAL Audit Report — {esc(source_filename)}</title>
<style>
:root{{
  --bg:#0D0F1A;--panel:#13162A;--card:#181C30;--brd:#252C45;
  --brdl:#303A58;--acc:#6374F8;--red:#E95555;--amb:#DCA032;
  --grn:#2EAD7A;--txt:#E2E8F0;--dim:#8C90A4;--mono:"Courier New",monospace
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--txt);font-family:var(--mono);font-size:13px;line-height:1.6}}
a{{color:var(--acc);text-decoration:none}}
.confidential{{
  background:#1a0e00;border-bottom:1px solid #92400e;color:#fbbf24;
  font-size:11px;font-weight:bold;letter-spacing:2px;text-align:center;
  padding:10px 40px;text-transform:uppercase
}}
header{{background:var(--panel);border-bottom:2px solid var(--brdl);padding:32px 48px 24px}}
.brand{{color:var(--acc);font-size:10px;font-weight:bold;letter-spacing:3px;margin-bottom:10px}}
header h1{{font-size:26px;font-weight:bold;color:#F8FAFC;margin-bottom:8px}}
.meta{{color:var(--dim);font-size:12px;line-height:2}}
.meta span{{color:var(--txt);margin-left:8px}}
main{{max-width:1280px;margin:0 auto;padding:36px 48px}}
section{{margin-bottom:48px}}
h2{{font-size:13px;font-weight:bold;letter-spacing:2px;color:var(--dim);
    text-transform:uppercase;border-bottom:1px solid var(--brd);
    padding-bottom:10px;margin-bottom:20px;display:flex;align-items:center;gap:12px}}
p.note{{color:var(--dim);font-size:11px;margin:8px 0 14px;font-style:italic}}
.stat-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:12px;margin-bottom:4px}}
.stat-card{{background:var(--card);border:1px solid var(--brdl);border-radius:8px;padding:18px 20px}}
.stat-val{{font-size:30px;font-weight:bold;line-height:1.1}}
.stat-label{{color:var(--dim);font-size:10px;margin-top:6px;text-transform:uppercase;letter-spacing:1px}}
.stat-sub{{color:var(--dim);font-size:11px;margin-top:2px}}
.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:32px}}
@media(max-width:900px){{.grid-2{{grid-template-columns:1fr}}}}
table{{width:100%;border-collapse:collapse;background:var(--card);
       border:1px solid var(--brd);font-size:12px}}
thead tr{{background:var(--panel)}}
th{{color:var(--dim);font-size:10px;font-weight:bold;letter-spacing:1px;
    text-transform:uppercase;padding:10px 14px;text-align:left;
    border-bottom:1px solid var(--brd)}}
td{{padding:7px 14px;border-bottom:1px solid var(--brd);vertical-align:top;word-break:break-word}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:rgba(99,116,248,0.06)}}
.bar-label{{width:200px;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.bar-cell{{width:100%;padding-right:12px}}
.bar-track{{height:6px;background:var(--brd);border-radius:3px;overflow:hidden}}
.bar-fill{{height:100%;border-radius:3px;transition:width .3s}}
.bar-count{{width:64px;text-align:right;font-weight:bold;white-space:nowrap;color:var(--txt)}}
.bar-badge{{width:100px;text-align:right;white-space:nowrap}}
.subj{{max-width:280px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.badge-crit{{background:#2d0010;color:#FF4D6D;border:1px solid #FF4D6D;
             padding:2px 10px;border-radius:3px;font-size:11px;font-weight:bold}}
.badge-high{{background:#2d1a00;color:#DCA032;border:1px solid #DCA032;
             padding:2px 10px;border-radius:3px;font-size:11px;font-weight:bold}}
footer{{border-top:1px solid var(--brd);padding:20px 48px;color:var(--dim);
        font-size:11px;text-align:center;margin-top:24px}}
.privacy{{background:#1a0e00;border-top:1px solid #92400e;color:#92400e;
          font-size:10px;padding:12px 48px;text-align:center}}
</style>
</head>
<body>
<div class="confidential">&#9888; Confidential — Contains Personal Information — Handle per Engagement Data Handling Policy &#9888;</div>
<header>
  <div class="brand">◎ UAL AUDIT LOG PARSER — INVESTIGATION REPORT</div>
  <h1>{esc(source_filename) or "UAL Export"}</h1>
  <div class="meta">
    <div>Source file<span>{esc(source_filename)}</span></div>
    <div>Generated<span>{esc(generated)}</span></div>
    <div>Total rows parsed<span>{total:,}</span></div>
  </div>
</header>
<main>

  <section>
    <h2>Summary Statistics</h2>
    <div class="stat-grid">{stats_html}</div>
  </section>

  {rule_section}
  {del_section}

  <section>
    <h2>Operation Breakdown</h2>
    <table>
      <thead><tr><th>Operation</th><th>Count</th><th></th><th>Severity</th></tr></thead>
      <tbody>{ops_rows}</tbody>
    </table>
  </section>

  <div class="grid-2">
    <section>
      <h2>Category Breakdown</h2>
      <table>
        <thead><tr><th>Category</th><th>Count</th><th></th><th></th></tr></thead>
        <tbody>{cat_rows}</tbody>
      </table>
    </section>
    <section>
      <h2>Result Status</h2>
      <table>
        <thead><tr><th>Status</th><th>Count</th><th></th><th></th></tr></thead>
        <tbody>{res_rows}</tbody>
      </table>
    </section>
  </div>

  <div class="grid-2">
    <section>
      <h2>Top Users by Event Count</h2>
      <table>
        <thead><tr><th>User ID</th><th>Count</th><th></th><th></th></tr></thead>
        <tbody>{usr_rows if usr_rows else '<tr><td colspan="4" style="color:var(--dim)">No user data</td></tr>'}</tbody>
      </table>
    </section>
    <section>
      <h2>Top Source IPs</h2>
      <table>
        <thead><tr><th>IP Address</th><th>Count</th><th></th><th></th></tr></thead>
        <tbody>{ip_rows if ip_rows else '<tr><td colspan="4" style="color:var(--dim)">No IP data</td></tr>'}</tbody>
      </table>
    </section>
  </div>

</main>
<div class="privacy">PRIVACY NOTICE — This report contains personal information derived from Microsoft 365 audit logs. Distribution must be limited to authorised personnel. Retain and dispose in accordance with applicable privacy law and your engagement data handling policy.</div>
<footer>UAL Audit Log Parser · Yuvi Kapoor · {esc(generated)} · {total:,} rows analysed</footer>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    return output_path



# ─── GUI ──────────────────────────────────────────────────────────────────────

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

MONO = "Courier New"

# ThreadHunter-derived theme dict — single source of truth for all colours
TC = {
    # Light canvas — dark chrome sidebar, light workspace
    "bg":                "#F5F6FA",   # main workspace canvas
    "sidebar":           "#1A1D27",   # deep navy sidebar (dark chrome)
    "panel":             "#FFFFFF",   # top/bottom bars — white
    "card":              "#FFFFFF",   # cards / panels
    "card2":             "#F5F6FA",   # alt card surface
    "border":            "#E4E6EF",   # hairline dividers
    "border_lt":         "#D0D3E8",   # slightly stronger border
    "accent":            "#6374F8",   # iris primary
    "accent2":           "#E95555",   # crimson error
    "accent3":           "#2EAD7A",   # emerald success
    "accent4":           "#DCA032",   # amber warning
    "text":              "#1A1D27",   # primary text on light bg
    "text_dim":          "#4E5263",   # secondary text
    "text_bright":       "#FFFFFF",   # text on dark surfaces (sidebar)
    "text_muted":        "#8C90A4",   # muted / hints
    "btn_primary":       "#6374F8",
    "btn_primary_hover": "#8B9EFF",
    "btn_secondary":     "#FFFFFF",
    "status_ok":         "#2EAD7A",
    "status_warn":       "#DCA032",
    "status_err":        "#E95555",
}


class UALParserApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("UAL Audit Log Parser  |  DFIR Tooling")
        self.geometry("1440x860")
        self.minsize(1100, 680)
        self.configure(fg_color="#F5F6FA")

        # State
        self.input_path   = tk.StringVar()
        self.output_dir   = tk.StringVar()
        self.output_fmt   = tk.StringVar(value="xlsx")
        self.status_var   = tk.StringVar(value="● READY")
        self.parsed_rows: list[dict] = []
        self.parse_errors: list      = []
        self.error_counts: dict      = {"critical": 0, "warn": 0, "info": 0}

        # Initialise ttk separator style before any widget uses it
        _s = ttk.Style()
        _s.theme_use("clam")
        _s.configure("UAL.TSeparator", background="#E4E6EF")
        self._apply_tree_style()

        self._build_ui()
        self.after(200, self._set_icon)

    def _set_icon(self):
        """Draw document+magnifier logo icon and apply via wm_iconphoto."""
        try:
            from PIL import Image, ImageDraw
            import tempfile, os
            S = 64
            img = Image.new("RGBA", (S, S), (0, 0, 0, 0))
            d = ImageDraw.Draw(img)
            IRIS = (99, 116, 248, 255)
            DIM  = (61,  74, 158, 255)
            BG   = (26,  29,  39, 255)
            # Document body
            d.rounded_rectangle([10, 8, 42, 52], radius=3, fill=BG, outline=IRIS, width=2)
            # Dog-ear
            d.polygon([(34, 8), (42, 16), (34, 16)], fill=BG, outline=IRIS)
            d.line([(34, 8), (42, 16)], fill=IRIS, width=2)
            d.line([(42, 16), (34, 16)], fill=IRIS, width=2)
            # Data rows
            for y, clr in [(22, IRIS), (29, DIM), (36, IRIS), (43, DIM)]:
                x2 = 32 if y in (22, 36) else 28
                d.line([(15, y), (x2, y)], fill=clr, width=2)
            # Magnifier
            d.ellipse([36, 38, 56, 58], outline=IRIS, width=3)
            d.line([(53, 55), (61, 63)], fill=IRIS, width=3)
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(tmp.name, "PNG")
            tmp.close()
            photo = tk.PhotoImage(file=tmp.name)
            self._icon_ref = photo
            self.wm_iconphoto(True, photo)
            os.unlink(tmp.name)
        except Exception:
            pass

    def _apply_tree_style(self):
        s = ttk.Style()
        s.configure("UAL.Treeview",
                     background="#FFFFFF",
                     foreground="#1A1D27",
                     fieldbackground="#FFFFFF",
                     rowheight=22,
                     font=(MONO, 10),
                     borderwidth=0,
                     relief="flat")
        s.configure("UAL.Treeview.Heading",
                     background="#F5F6FA",
                     foreground=TC["text_dim"],
                     relief="flat",
                     borderwidth=0,
                     font=(MONO, 9, "bold"))
        s.map("UAL.Treeview",
              background=[("selected", "#4A57D4")],
              foreground=[("selected", "#FFFFFF")])

    # ── Full UI ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        tc = TC
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=0)   # footer fixed
        self.grid_columnconfigure(1, weight=1)

        self._build_sidebar()
        self._build_main()
        self._build_footer()

    # ── Footer ────────────────────────────────────────────────────────────────

    def _build_footer(self):
        tc = TC
        bar = ctk.CTkFrame(self, fg_color=tc["panel"], corner_radius=0, height=36)
        bar.grid(row=1, column=0, columnspan=2, sticky="ew")
        bar.grid_propagate(False)
        bar.grid_columnconfigure(0, weight=1)

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(inner,
                     text="UAL Operations Parser  —  Created by Yuvi Kapoor  —  ",
                     font=ctk.CTkFont(family=MONO, size=12),
                     text_color="#8C90A4").pack(side="left")

        lnk = ctk.CTkLabel(inner,
                            text="linkedin.com/in/yuvi-kapoor-5a38521a5",
                            font=ctk.CTkFont(family=MONO, size=12),
                            text_color=tc["accent"],
                            cursor="hand2")
        lnk.pack(side="left")
        lnk.bind("<Button-1>", lambda e: __import__("webbrowser").open(
            "https://www.linkedin.com/in/yuvi-kapoor-5a38521a5"))
        lnk.bind("<Enter>", lambda e: lnk.configure(text_color=tc["btn_primary_hover"]))
        lnk.bind("<Leave>", lambda e: lnk.configure(text_color=tc["accent"]))

    # ── Sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self):
        tc = TC

        sidebar = ctk.CTkFrame(self, fg_color=tc["sidebar"],
                                corner_radius=0, width=290)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        sidebar.grid_columnconfigure(0, weight=1)
        sidebar.grid_rowconfigure(2, weight=1)   # scroll area expands

        # ── Brand header — pinned ─────────────────────────────────────────────
        brand = ctk.CTkFrame(sidebar, fg_color=tc["sidebar"], corner_radius=0)
        brand.grid(row=0, column=0, sticky="ew")
        brand.grid_columnconfigure(0, weight=1)

        bi = ctk.CTkFrame(brand, fg_color="transparent")
        bi.pack(fill="x", padx=18, pady=(14, 12))

        # Logo canvas — document + magnifier drawn in Iris blue
        logo_canvas = tk.Canvas(bi, width=48, height=48,
                                bg=tc["sidebar"], highlightthickness=0)
        logo_canvas.pack(anchor="w", pady=(0, 8))
        # Document body
        logo_canvas.create_rectangle(4, 2, 30, 42, outline="#6374F8",
                                     width=2, fill="#1A1D27")
        # Dog-ear
        logo_canvas.create_polygon(24, 2, 30, 8, 30, 2,
                                   outline="#6374F8", fill="#1A1D27", width=1)
        logo_canvas.create_line(24, 2, 30, 8, fill="#6374F8", width=1)
        logo_canvas.create_line(24, 8, 30, 8, fill="#6374F8", width=1)
        # Data rows
        logo_canvas.create_line(8, 16, 22, 16, fill="#6374F8", width=2)
        logo_canvas.create_line(8, 22, 19, 22, fill="#3D4A9E", width=1)
        logo_canvas.create_line(8, 28, 23, 28, fill="#6374F8", width=2)
        logo_canvas.create_line(8, 34, 17, 34, fill="#3D4A9E", width=1)
        # Magnifier circle
        logo_canvas.create_oval(27, 28, 43, 44, outline="#6374F8", width=2)
        # Magnifier handle
        logo_canvas.create_line(40, 41, 48, 48, fill="#6374F8", width=2)

        ctk.CTkLabel(bi, text="UAL OPERATIONS PARSER",
                     font=ctk.CTkFont(family=MONO, size=12, weight="bold"),
                     text_color=tc["text_bright"], anchor="w").pack(fill="x")
        ctk.CTkLabel(bi, text="M365 BEC AUDIT INTELLIGENCE",
                     font=ctk.CTkFont(family=MONO, size=9),
                     text_color=tc["accent"], anchor="w").pack(fill="x")

        ttk.Separator(sidebar, orient="horizontal",
                      style="UAL.TSeparator").grid(row=1, column=0, sticky="ew")

        # ── Scrollable content ────────────────────────────────────────────────
        scroll = ctk.CTkScrollableFrame(
            sidebar,
            fg_color=tc["sidebar"],
            scrollbar_button_color="#383E5E",
            scrollbar_button_hover_color=tc["accent"],
            corner_radius=0,
        )
        scroll.grid(row=2, column=0, sticky="nsew")
        scroll.grid_columnconfigure(0, weight=1)

        ttk.Separator(sidebar, orient="horizontal",
                      style="UAL.TSeparator").grid(row=3, column=0, sticky="ew")

        # ── Bottom pinned controls ────────────────────────────────────────────
        bottom = ctk.CTkFrame(sidebar, fg_color="#13162A", corner_radius=0)
        bottom.grid(row=4, column=0, sticky="ew")
        bottom.grid_columnconfigure(0, weight=1)

        self.parse_btn = ctk.CTkButton(
            bottom, text="▶  PARSE LOG",
            fg_color=tc["btn_primary"], hover_color=tc["btn_primary_hover"],
            text_color="#FFFFFF", height=40, corner_radius=4,
            font=ctk.CTkFont(family=MONO, size=12, weight="bold"),
            command=self._run_parse)
        self.parse_btn.grid(row=0, column=0, padx=10, pady=(10, 4), sticky="ew")

        self.export_btn = ctk.CTkButton(
            bottom, text="⬇  EXPORT OUTPUT",
            fg_color="transparent", hover_color="#2A2F4A",
            text_color="#8C90A4", border_color="#383E5E", border_width=1,
            height=32, corner_radius=4,
            font=ctk.CTkFont(family=MONO, size=11),
            command=self._export, state="disabled")
        self.export_btn.grid(row=1, column=0, padx=10, pady=(0, 4), sticky="ew")

        ctk.CTkButton(
            bottom, text="✕  CLEAR",
            fg_color="transparent", hover_color="#2A2F4A",
            text_color="#6C7185", border_color="#2A2F4A", border_width=1,
            height=26, corner_radius=4,
            font=ctk.CTkFont(family=MONO, size=10),
            command=self._clear).grid(row=2, column=0, padx=10, pady=(0, 10), sticky="ew")

        # ── Helpers (scoped to scroll frame) ─────────────────────────────────
        def _section_div(text):
            """ThreadHunter-style ── LABEL ──────────── divider."""
            outer = ctk.CTkFrame(scroll, fg_color="transparent")
            outer.pack(fill="x", padx=14, pady=(14, 5))
            ctk.CTkFrame(outer, fg_color="#383E5E",
                         height=1, corner_radius=0, width=10).pack(
                side="left", ipadx=0)
            ctk.CTkLabel(outer,
                         text=f"  {text}  ",
                         font=ctk.CTkFont(family=MONO, size=11, weight="bold"),
                         text_color="#8C90A4").pack(side="left")
            ctk.CTkFrame(outer, fg_color=tc["border_lt"],
                         height=1, corner_radius=0).pack(
                side="left", fill="x", expand=True)

        def _card():
            """Elevated card panel — slightly lighter than sidebar bg."""
            c = ctk.CTkFrame(scroll, fg_color="#21253A",
                              corner_radius=8, border_width=1,
                              border_color="#383E5E")
            c.pack(fill="x", padx=10, pady=(0, 6))
            return c

        # ── INPUT FILE ────────────────────────────────────────────────────────
        _section_div("INPUT FILE")
        c = _card()

        self._ual_lbl = ctk.CTkLabel(
            c, text="No file selected",
            font=ctk.CTkFont(family=MONO, size=11),
            text_color="#8C90A4", wraplength=240, justify="left", anchor="w")
        self._ual_lbl.pack(fill="x", padx=12, pady=(10, 4))

        file_row = ctk.CTkFrame(c, fg_color="transparent")
        file_row.pack(fill="x", padx=8, pady=(0, 8))
        file_row.grid_columnconfigure(0, weight=1)
        self._input_entry = ctk.CTkEntry(
            file_row, textvariable=self.input_path,
            placeholder_text="Select UAL export (.csv / .xlsx)",
            fg_color="#13162A", border_color="#383E5E", border_width=1,
            text_color="#FFFFFF", placeholder_text_color="#6C7185",
            font=ctk.CTkFont(family=MONO, size=9), height=28, corner_radius=3)
        self._input_entry.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ctk.CTkButton(
            file_row, text="›", width=28, height=28,
            fg_color=tc["btn_primary"], hover_color=tc["btn_primary_hover"],
            text_color="white", corner_radius=3,
            font=ctk.CTkFont(family=MONO, size=13, weight="bold"),
            command=self._browse_input).grid(row=0, column=1)

        # ── OUTPUT DIRECTORY ─────────────────────────────────────────────────
        _section_div("OUTPUT DIRECTORY")
        c = _card()

        self._out_lbl = ctk.CTkLabel(
            c, text="Same folder as input",
            font=ctk.CTkFont(family=MONO, size=11),
            text_color="#8C90A4", wraplength=240, anchor="w")
        self._out_lbl.pack(fill="x", padx=12, pady=(10, 4))

        out_row = ctk.CTkFrame(c, fg_color="transparent")
        out_row.pack(fill="x", padx=8, pady=(0, 8))
        out_row.grid_columnconfigure(0, weight=1)
        self._output_entry = ctk.CTkEntry(
            out_row, textvariable=self.output_dir,
            placeholder_text="Default: same as input file",
            fg_color="#13162A", border_color="#383E5E", border_width=1,
            text_color="#FFFFFF", placeholder_text_color="#6C7185",
            font=ctk.CTkFont(family=MONO, size=9), height=28, corner_radius=3)
        self._output_entry.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ctk.CTkButton(
            out_row, text="›", width=28, height=28,
            fg_color=tc["btn_primary"], hover_color=tc["btn_primary_hover"],
            text_color="white", corner_radius=3,
            font=ctk.CTkFont(family=MONO, size=13, weight="bold"),
            command=self._browse_output).grid(row=0, column=1)

        # ── OUTPUT FORMAT ─────────────────────────────────────────────────────
        _section_div("OUTPUT FORMAT")
        c = _card()
        fmt_row = ctk.CTkFrame(c, fg_color="transparent")
        fmt_row.pack(fill="x", padx=12, pady=10)
        ctk.CTkLabel(fmt_row, text="FORMAT",
                     font=ctk.CTkFont(family=MONO, size=10, weight="bold"),
                     text_color="#8C90A4").pack(side="left", padx=(0, 10))
        ctk.CTkRadioButton(
            fmt_row, text="XLSX", variable=self.output_fmt, value="xlsx",
            text_color="#FFFFFF", font=ctk.CTkFont(family=MONO, size=10),
            fg_color="#6374F8", hover_color="#8B9EFF"
        ).pack(side="left", padx=(0, 14))
        ctk.CTkRadioButton(
            fmt_row, text="CSV", variable=self.output_fmt, value="csv",
            text_color="#FFFFFF", font=ctk.CTkFont(family=MONO, size=10),
            fg_color="#6374F8", hover_color="#8B9EFF"
        ).pack(side="left")

        # ── OPERATION FILTER ──────────────────────────────────────────────────
        _section_div("OPERATION FILTER")
        c = _card()

        filters = [
            ("email_send",  "Email – Create / Send / SendAs"),
            ("delete",      "Delete – Hard / Soft / Purge"),
            ("mail_access", "MailItemsAccessed"),
            ("file_folder", "File / Folder / SharePoint"),
            ("inbox_rule",  "Inbox Rules"),
            ("teams",       "Teams – Channels / Members / Apps"),
            ("sign_in",     "Sign-In / Logon"),
            ("admin",       "Admin / Config Changes"),
            ("generic",     "Other / Unclassified"),
        ]
        self.filter_vars = {}
        for key, label in filters:
            var = tk.BooleanVar(value=True)
            self.filter_vars[key] = var
            ctk.CTkCheckBox(
                c, text=label, variable=var,
                text_color="#FFFFFF", font=ctk.CTkFont(family=MONO, size=10),
                fg_color="#6374F8", hover_color="#8B9EFF",
                checkmark_color="white", border_color="#383E5E",
                height=24
            ).pack(anchor="w", padx=12, pady=2)
        ctk.CTkFrame(c, fg_color="transparent", height=6).pack()

        # Spacer at bottom of scroll area
        ctk.CTkFrame(scroll, fg_color="transparent", height=16).pack()

    # ── Main content ──────────────────────────────────────────────────────────

    def _build_main(self):
        tc = TC
        main = ctk.CTkFrame(self, fg_color=tc["bg"], corner_radius=0)
        main.grid(row=0, column=1, sticky="nsew")
        # row 0=stats bar, 1=sep, 2=log+progress, 3=sep, 4=preview (expands)
        main.grid_rowconfigure(4, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # ── Stats bar ─────────────────────────────────────────────────────────
        stats_bar = ctk.CTkFrame(main, fg_color="#FFFFFF",
                                  corner_radius=0, height=60)
        stats_bar.grid(row=0, column=0, sticky="ew")
        stats_bar.grid_propagate(False)
        stats_bar.grid_columnconfigure(tuple(range(6)), weight=1)

        self.stat_labels = {}
        stats = [
            ("total",       "TOTAL ROWS",    "─"),
            ("parsed",      "PARSED",        "─"),
            ("email_send",  "EMAIL OPS",     "─"),
            ("delete",      "DELETES",       "─"),
            ("mail_access", "MAIL ACCESS",   "─"),
            ("inbox_rule",  "INBOX RULES",   "─"),
        ]
        for col, (key, label, default) in enumerate(stats):
            box = ctk.CTkFrame(stats_bar, fg_color="transparent")
            box.grid(row=0, column=col, padx=4, pady=8, sticky="nsew")
            ctk.CTkLabel(box, text=label,
                         font=ctk.CTkFont(family=MONO, size=8),
                         text_color="#8C90A4").pack()
            lbl = ctk.CTkLabel(box, text=default,
                               font=ctk.CTkFont(family=MONO, size=16, weight="bold"),
                               text_color=tc["accent"])
            lbl.pack()
            self.stat_labels[key] = lbl

        # ── Separator ─────────────────────────────────────────────────────────
        ttk.Separator(main, orient="horizontal",
                      style="UAL.TSeparator").grid(row=1, column=0, sticky="ew")

        # ── Log + progress panel ──────────────────────────────────────────────
        log_outer = ctk.CTkFrame(main, fg_color="#FFFFFF",
                                  corner_radius=0, height=180, border_width=0)
        log_outer.grid(row=2, column=0, sticky="ew")
        log_outer.grid_propagate(False)
        log_outer.grid_rowconfigure(1, weight=1)
        log_outer.grid_columnconfigure(0, weight=1)

        # Log header bar (like ThreadHunter's ACTIVITY LOG strip)
        log_hdr = ctk.CTkFrame(log_outer, fg_color="#F5F6FA",
                                corner_radius=0, height=36)
        log_hdr.grid(row=0, column=0, sticky="ew")
        log_hdr.grid_propagate(False)
        log_hdr.grid_columnconfigure(0, weight=1)

        hdr_inner = ctk.CTkFrame(log_hdr, fg_color="transparent")
        hdr_inner.grid(row=0, column=0, sticky="ew", padx=14, pady=6)
        hdr_inner.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr_inner, text="ACTIVITY LOG",
                     font=ctk.CTkFont(family=MONO, size=12, weight="bold"),
                     text_color="#4E5263").grid(row=0, column=0, sticky="w")

        # Progress bar in header — matches ThreadHunter's 3px strip
        self.progress = ctk.CTkProgressBar(
            hdr_inner, mode="determinate",
            fg_color=tc["border"], progress_color=tc["accent"], height=3)
        self.progress.grid(row=0, column=1, padx=(8, 8), sticky="ew")
        self.progress.set(0)

        ctk.CTkButton(hdr_inner, text="CLEAR",
                      width=64, height=26,
                      fg_color="transparent", hover_color="#FDECEA",
                      text_color="#8C90A4",
                      border_color="#E4E6EF", border_width=1,
                      font=ctk.CTkFont(family=MONO, size=11, weight="bold"),
                      command=self._clear_log).grid(row=0, column=2, sticky="e")

        # Separator under log header
        ctk.CTkFrame(log_outer, fg_color="#E4E6EF",
                     height=1, corner_radius=0).grid(row=0, column=0, sticky="sew")

        self.log_box = ctk.CTkTextbox(
            log_outer,
            fg_color="#FAFBFD",
            text_color="#4E5263",
            font=ctk.CTkFont(family=MONO, size=11),
            state="disabled")
        self.log_box.grid(row=1, column=0, padx=4, pady=(0, 4), sticky="nsew")

        # Status label below log (● READY pattern from ThreadHunter)
        status_bar = ctk.CTkFrame(log_outer, fg_color="#F5F6FA",
                                   corner_radius=0, height=32)
        status_bar.grid(row=2, column=0, sticky="ew")
        status_bar.grid_propagate(False)

        ctk.CTkLabel(status_bar,
                     textvariable=self.status_var,
                     font=ctk.CTkFont(family=MONO, size=11),
                     text_color=tc["status_ok"],
                     anchor="w").pack(side="left", padx=14, pady=6)

        # ── Separator above preview table ─────────────────────────────────────
        ttk.Separator(main, orient="horizontal",
                      style="UAL.TSeparator").grid(row=3, column=0, sticky="ew")

        # ── Preview table ─────────────────────────────────────────────────────
        preview_outer = ctk.CTkFrame(main, fg_color="#FFFFFF", corner_radius=0)
        preview_outer.grid(row=4, column=0, sticky="nsew")
        preview_outer.grid_rowconfigure(1, weight=1)
        preview_outer.grid_columnconfigure(0, weight=1)

        # Table header bar
        tbl_hdr = ctk.CTkFrame(preview_outer, fg_color="#F5F6FA",
                                corner_radius=0, height=26)
        tbl_hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        tbl_hdr.grid_propagate(False)
        tbl_hdr.grid_columnconfigure(0, weight=1)

        tbl_inner = ctk.CTkFrame(tbl_hdr, fg_color="transparent")
        tbl_inner.grid(row=0, column=0, sticky="ew", padx=10, pady=3)
        tbl_inner.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(tbl_inner,
                     text="PARSED PREVIEW  (first 200 rows)",
                     font=ctk.CTkFont(family=MONO, size=11, weight="bold"),
                     text_color="#4E5263").grid(row=0, column=0, sticky="w")

        ctk.CTkFrame(preview_outer, fg_color="#E4E6EF",
                     height=1, corner_radius=0).grid(row=0, column=0,
                                                       columnspan=2, sticky="sew")

        # Treeview with proper scrollbars
        tree_frame = ctk.CTkFrame(preview_outer, fg_color="transparent",
                                   corner_radius=0)
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        preview_cols = ["Row", "CreationTime", "Operation", "OperationCategory",
                        "UserId", "Subject", "Path/Folder", "ResultStatus"]

        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")

        self.tree = ttk.Treeview(
            tree_frame, columns=preview_cols, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set,
            selectmode="browse", style="UAL.Treeview"
        )
        vsb.configure(command=self.tree.yview)
        hsb.configure(command=self.tree.xview)

        col_w = {"Row": 44, "CreationTime": 210, "Operation": 130,
                  "OperationCategory": 120, "UserId": 200,
                  "Subject": 240, "Path/Folder": 200, "ResultStatus": 90}
        for c in preview_cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_w.get(c, 120), anchor="w", stretch=False)

        # Row tags — light mode tinted rows
        cat_bg = {
            "email_send":          "#EEF5EE",
            "delete":              "#FDECEA",
            "mail_access":         "#F2EEFB",
            "file_folder":         "#EEF3FB",
            "inbox_rule":          "#FDF6E3",
            "teams":               "#EEF5FB",
            "sign_in":             "#EEF8F4",
            "admin":               "#FDE8E8",
            "generic":             "#FFFFFF",
            "parse_error_critical":"#FDECEA",
            "parse_error_warn":    "#FDF6E3",
        }
        cat_fg = {
            "parse_error_critical": "#C0392B",
            "parse_error_warn":     "#9A6700",
        }
        for cat, bg in cat_bg.items():
            self.tree.tag_configure(cat, background=bg,
                                    foreground=cat_fg.get(cat, "#1A1D27"))

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Keyboard shortcut bar (ThreadHunter-style)
        self._build_shortcut_bar(main, row=5)

    def _build_shortcut_bar(self, parent, row: int):
        tc = TC
        bar = ctk.CTkFrame(parent, fg_color="#F5F6FA",
                            corner_radius=0, height=22)
        bar.grid(row=row, column=0, sticky="ew")
        bar.grid_propagate(False)

        shortcuts = [
            ("Enter", "Parse"),
            ("Ctrl+E", "Export"),
            ("Ctrl+O", "Open File"),
            ("Ctrl+L", "Clear Log"),
        ]
        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.place(relx=0, rely=0.5, anchor="w", x=10)
        for i, (key, label) in enumerate(shortcuts):
            f = ctk.CTkFrame(inner, fg_color="transparent")
            f.grid(row=0, column=i * 2)
            ctk.CTkLabel(f, text=key,
                         font=ctk.CTkFont(family=MONO, size=8, weight="bold"),
                         text_color="#6374F8").grid(row=0, column=0)
            ctk.CTkLabel(f, text=f" {label}",
                         font=ctk.CTkFont(family=MONO, size=8),
                         text_color="#8C90A4").grid(row=0, column=1)
            if i < len(shortcuts) - 1:
                ctk.CTkLabel(inner, text="  ·  ",
                             font=ctk.CTkFont(family=MONO, size=8),
                             text_color=tc["text_muted"]).grid(row=0, column=i * 2 + 1)

        # Bind shortcuts
        self.bind_all("<Control-e>", lambda e: self._export())
        self.bind_all("<Control-o>", lambda e: self._browse_input())
        self.bind_all("<Control-l>", lambda e: self._clear_log())
        self.bind_all("<Return>",    lambda e: self._run_parse() if self.parse_btn.cget("state") == "normal" else None)

    # ── Actions ───────────────────────────────────────────────────────────────

    def _browse_input(self):
        p = filedialog.askopenfilename(
            title="Select UAL Export",
            filetypes=[("Audit Log Files", "*.csv *.xlsx *.xls"), ("All Files", "*.*")]
        )
        if p:
            self.input_path.set(p)
            self._ual_lbl.configure(
                text=f"✔  {Path(p).name}", text_color="#2EAD7A")

    def _browse_output(self):
        d = filedialog.askdirectory(title="Select Output Directory")
        if d:
            self.output_dir.set(d)
            self._out_lbl.configure(
                text=f"✔  {d}", text_color="#2EAD7A")

    def _log(self, msg: str, level: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        prefix = {"info": "  ", "ok": "✔ ", "warn": "△ ", "err": "✗ "}.get(level, "  ")
        line = f"[{ts}] {prefix}{msg}\n"
        self.log_box.configure(state="normal")
        self.log_box.insert("end", line)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _set_status(self, msg: str, level: str = "ok"):
        col = {"ok": TC["status_ok"], "warn": TC["status_warn"],
               "err": TC["status_err"]}.get(level, TC["text_dim"])
        self.status_var.set(msg)
        # Find and recolour the status label
        try:
            main_frame = self.grid_slaves(row=0, column=1)[0]
            for widget in main_frame.winfo_children():
                if hasattr(widget, "grid_info") and widget.grid_info().get("row") == 2:
                    for sub in widget.winfo_children():
                        for lbl in sub.winfo_children() if hasattr(sub, "winfo_children") else []:
                            if hasattr(lbl, "cget") and lbl.cget("textvariable") == str(self.status_var):
                                lbl.configure(text_color=col)
        except Exception:
            pass

    def _clear(self):
        self.input_path.set("")
        self.output_dir.set("")
        self.parsed_rows = []
        self.parse_errors = []
        self.error_counts = {"critical": 0, "warn": 0, "info": 0}
        self._clear_log()
        self.progress.set(0)
        self.export_btn.configure(state="disabled", text="⬇  EXPORT OUTPUT")
        for item in self.tree.get_children():
            self.tree.delete(item)
        for key in self.stat_labels:
            self.stat_labels[key].configure(text="─")
        self._ual_lbl.configure(text="No file selected", text_color=TC["text_dim"])
        self._out_lbl.configure(text="Same folder as input", text_color=TC["text_dim"])
        self.status_var.set("● READY")

    # ── Parse ─────────────────────────────────────────────────────────────────

    def _run_parse(self):
        inp = self.input_path.get().strip()
        if not inp:
            messagebox.showwarning("No Input", "Select a UAL export file first.")
            return
        if not os.path.isfile(inp):
            messagebox.showerror("File Not Found", f"Cannot find:\n{inp}")
            return
        self.parse_btn.configure(state="disabled", text="Parsing…")
        self.export_btn.configure(state="disabled")
        self.progress.set(0)
        self.status_var.set("● Running…")
        threading.Thread(target=self._parse_worker, args=(inp,), daemon=True).start()

    def _parse_worker(self, filepath: str):
        try:
            self._log(f"Loading: {Path(filepath).name}")
            raw_rows = load_input_file(filepath)
            total = len(raw_rows)
            self._log(f"Loaded {total:,} rows.")
            self.after(0, lambda: self.stat_labels["total"].configure(text=f"{total:,}"))

            enabled_cats = {k for k, v in self.filter_vars.items() if v.get()}
            parsed = []
            all_errors = []
            critical_count = 0
            warn_count = 0

            for i, row in enumerate(raw_rows):
                pr, errs = parse_row(i + 1, row)
                all_errors.extend(errs)
                for e in errs:
                    if e.severity == ERR_CRITICAL:
                        critical_count += 1
                        self._log(
                            f"Row {i+1} [{pr.get('Operation','')}] CRITICAL: {e.error_type} — {e.detail[:80]}",
                            "err")
                    elif e.severity == ERR_WARN:
                        warn_count += 1
                        if warn_count <= 20:
                            self._log(
                                f"Row {i+1} [{pr.get('Operation','')}] WARN: {e.error_type} — {e.detail[:80]}",
                                "warn")

                if pr["OperationCategory"] in enabled_cats:
                    parsed.append(pr)
                if i % 200 == 0:
                    pct = i / max(total, 1)
                    self.after(0, lambda p=pct: self.progress.set(p))

            if warn_count > 20:
                self._log(f"… and {warn_count - 20} more WARNING rows. See error output.", "warn")

            self.parsed_rows = parsed
            self.parse_errors = all_errors
            self.error_counts = {"critical": critical_count, "warn": warn_count,
                                  "info": sum(1 for e in all_errors if e.severity == ERR_INFO)}
            self.after(0, self._parse_complete)
        except Exception as e:
            import traceback
            tb_str = traceback.format_exc()
            self.after(0, lambda msg=f"{e}\n\n{tb_str}": self._parse_error(msg))

    def _parse_complete(self):
        rows = self.parsed_rows
        ec   = self.error_counts
        self.progress.set(1.0)

        if ec["critical"] > 0:
            self._log(
                f"⚠  {ec['critical']} CRITICAL parse error(s) — rows flagged in output. "
                f"Review '⚠ Parse Errors' sheet or _ERRORS.csv.", "err")
        if ec["warn"] > 0:
            self._log(
                f"△  {ec['warn']} WARNING(s) — required fields empty. Schema variants may need handling.", "warn")
        if ec["info"] > 0:
            self._log(f"   {ec['info']} INFO notice(s) — unknown operations routed to generic.")

        total_errors = ec["critical"] + ec["warn"]
        if total_errors == 0:
            self._log(f"Parsing complete. {len(rows):,} rows processed. No errors.", "ok")
        else:
            self._log(f"Parsing complete. {len(rows):,} rows | {total_errors} error(s) — see error output.", "warn")

        from collections import Counter
        cats = Counter(r["OperationCategory"] for r in rows)
        self.stat_labels["parsed"].configure(text=f"{len(rows):,}")
        self.stat_labels["email_send"].configure(text=f"{cats.get('email_send', 0):,}")
        self.stat_labels["delete"].configure(
            text=f"{cats.get('delete', 0):,}")
        self.stat_labels["mail_access"].configure(
            text=f"{cats.get('mail_access', 0):,}")
        self.stat_labels["inbox_rule"].configure(
            text=f"{cats.get('inbox_rule', 0):,}")

        # Colour code high-risk stats
        del_count = cats.get("delete", 0)
        rule_count = cats.get("inbox_rule", 0)
        if del_count > 0:
            self.stat_labels["delete"].configure(text_color=TC["status_err"])
        if rule_count > 0:
            self.stat_labels["inbox_rule"].configure(text_color=TC["status_err"])

        # Populate preview
        for item in self.tree.get_children():
            self.tree.delete(item)
        preview_cols = ["Row", "CreationTime", "Operation", "OperationCategory",
                        "UserId", "Subject", "Path/Folder", "ResultStatus"]
        for r in rows[:200]:
            row_errors = r.get("_parse_errors", [])
            has_critical = any(e.severity == ERR_CRITICAL for e in row_errors)
            has_warn     = any(e.severity == ERR_WARN     for e in row_errors)
            tag = ("parse_error_critical" if has_critical else
                   "parse_error_warn"     if has_warn     else r["OperationCategory"])
            vals = tuple(str(r.get(c, "")) for c in preview_cols)
            self.tree.insert("", "end", values=vals, tags=(tag,))

        self.parse_btn.configure(state="normal", text="▶  PARSE LOG")
        self.export_btn.configure(state="normal")
        status_suffix = f" | ⚠ {total_errors} error(s)" if total_errors else " | ✔ Clean"
        self.status_var.set(f"● {len(rows):,} ROWS PARSED{status_suffix}")

    def _parse_error(self, msg: str):
        self._log(f"Error: {msg}", "err")
        self.parse_btn.configure(state="normal", text="▶  PARSE LOG")
        self.status_var.set("● Error — see log")
        messagebox.showerror("Parse Error", msg[:600])

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self):
        if not self.parsed_rows:
            messagebox.showwarning("Nothing to Export", "Run Parse first.")
            return

        inp = self.input_path.get().strip()
        stem = Path(inp).stem if inp else "ual_export"

        out_dir_raw = self.output_dir.get().strip()
        if out_dir_raw:
            out_dir = out_dir_raw
        elif inp:
            out_dir = str(Path(inp).parent)
        else:
            out_dir = str(Path.cwd())

        out_dir = os.path.normpath(out_dir)

        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception as e:
            msg = f"Cannot create output directory:\n{out_dir}\n\n{e}"
            self._log(f"Export error: {msg}", "err")
            messagebox.showerror("Invalid Output Directory", msg)
            return

        if not os.access(out_dir, os.W_OK):
            msg = f"Output directory is not writable:\n{out_dir}"
            self._log(f"Export error: {msg}", "err")
            messagebox.showerror("Permission Denied", msg)
            return

        fmt = self.output_fmt.get()
        if fmt == "xlsx" and not HAS_OPENPYXL:
            messagebox.showerror("Missing Dependency", "openpyxl not installed. Use CSV.")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{stem}_parsed_{ts}.{fmt}"
        out_path = os.path.join(out_dir, filename)

        self.export_btn.configure(state="disabled", text="Exporting…")
        self.progress.set(0)
        self._log(f"Writing to: {out_path}")
        self.status_var.set("● Exporting…")

        threading.Thread(
            target=self._export_worker,
            args=(out_path, fmt, out_dir),
            daemon=True
        ).start()

    def _export_worker(self, out_path: str, fmt: str, out_dir: str):
        try:
            if fmt == "xlsx":
                _, err_count = write_output_xlsx(out_path, self.parsed_rows)
            else:
                _, err_count = write_output_csv(out_path, self.parsed_rows)

            if not os.path.isfile(out_path):
                raise FileNotFoundError(f"File not found after write:\n{out_path}")

            # Generate HTML report alongside the data export
            stem = Path(out_path).stem
            html_path = os.path.join(out_dir, f"{stem}_report.html")
            source_name = Path(self.input_path.get()).name if self.input_path.get() else ""
            try:
                generate_html_report(self.parsed_rows, html_path, source_name)
            except Exception as _he:
                html_path = None
                self._log(f"HTML report error: {_he}", "warn")

            self.after(0, lambda hp=html_path: self._export_complete(
                out_path, fmt, out_dir, err_count, hp))

        except Exception as e:
            import traceback
            detail = traceback.format_exc()
            self.after(0, lambda err=str(e), tb=detail: self._export_failed(out_path, err, tb))

    def _export_complete(self, out_path: str, fmt: str, out_dir: str,
                          err_count: int, html_path: str = None):
        self.progress.set(1.0)
        self.export_btn.configure(state="normal", text="⬇  EXPORT OUTPUT")
        filename = Path(out_path).name

        err_note = ""
        if err_count > 0:
            sheet = "'⚠ Parse Errors' sheet" if fmt == "xlsx" else "companion _ERRORS.csv"
            err_note = f"\n\n⚠  {err_count} parse error(s) in {sheet}."
            self._log(f"Export complete with {err_count} error(s) flagged.", "warn")
        else:
            self._log("Export complete. No parse errors.", "ok")

        self._log(f"Exported → {out_path}", "ok")
        if html_path and os.path.isfile(html_path):
            self._log(f"Report  → {html_path}", "ok")

        self.status_var.set(f"● Exported: {filename}" + (f" | ⚠ {err_count} errors" if err_count else ""))

        # Build completion message and offer to open HTML report
        msg_lines = [f"Data export:\n  {out_path}"]
        if html_path and os.path.isfile(html_path):
            msg_lines.append(f"\nHTML report:\n  {Path(html_path).name}")
        if err_note:
            msg_lines.append(err_note)
        msg_lines.append("\nOpen HTML report in browser?")

        import subprocess, sys as _sys
        if html_path and os.path.isfile(html_path):
            if messagebox.askyesno("Export Complete", "\n".join(msg_lines)):
                try:
                    if _sys.platform == "win32":
                        subprocess.Popen(["start", "", html_path], shell=True)
                    elif _sys.platform == "darwin":
                        subprocess.Popen(["open", html_path])
                    else:
                        subprocess.Popen(["xdg-open", html_path])
                except Exception:
                    pass
        else:
            if messagebox.askyesno("Export Complete",
                                   f"Saved:\n{out_path}{err_note}\n\nOpen containing folder?"):
                try:
                    if _sys.platform == "win32":
                        subprocess.Popen(["explorer", "/select,", out_path])
                    elif _sys.platform == "darwin":
                        subprocess.Popen(["open", "-R", out_path])
                    else:
                        subprocess.Popen(["xdg-open", out_dir])
                except Exception:
                    pass



    def _export_failed(self, out_path: str, error: str, traceback_str: str):
        self.progress.set(0)
        self.export_btn.configure(state="normal", text="⬇  EXPORT OUTPUT")
        self._log(f"Export error: {error}", "err")
        self._log(traceback_str, "err")
        path_str = out_path or "(path not resolved)"
        messagebox.showerror("Export Failed",
                             f"Could not write:\n{path_str}\n\nError: {error}\n\n"
                             "Check the activity log for full details.")


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = UALParserApp()
    app.mainloop()
